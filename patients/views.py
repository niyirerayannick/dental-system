from io import BytesIO

from django.contrib import messages
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from accounts.models import User
from accounts.permissions import role_required
from appointments.models import Appointment
from treatments.models import Treatment
from .forms import PatientForm
from .models import PatientProfile


def calculate_age(date_of_birth):
    if not date_of_birth:
        return "N/A"
    today = timezone.localdate()
    return today.year - date_of_birth.year - ((today.month, today.day) < (date_of_birth.month, date_of_birth.day))


def base_patient_queryset(request):
    patients = PatientProfile.objects.select_related("user").prefetch_related("appointments", "treatments")
    search = request.GET.get("q", "").strip()
    gender = request.GET.get("gender", "").strip()
    status = request.GET.get("status", "").strip()
    appointment_status = request.GET.get("appointment_status", "").strip()

    if search:
        patients = patients.filter(
            Q(user__first_name__icontains=search)
            | Q(user__last_name__icontains=search)
            | Q(user__email__icontains=search)
            | Q(user__phone__icontains=search)
        )
    if gender:
        patients = patients.filter(gender=gender)
    if status == "active":
        patients = patients.filter(user__is_active=True)
    elif status == "inactive":
        patients = patients.filter(user__is_active=False)
    if appointment_status:
        patients = patients.filter(appointments__status=appointment_status)

    return patients.distinct()


def build_patient_rows(patients):
    rows = []
    for patient in patients:
        completed_appointments = [
            appointment for appointment in patient.appointments.all() if appointment.status == Appointment.Status.COMPLETED
        ]
        upcoming_appointments = [
            appointment
            for appointment in patient.appointments.all()
            if appointment.status in {Appointment.Status.PENDING, Appointment.Status.APPROVED}
            and appointment.appointment_date >= timezone.localdate()
        ]
        last_visit = max(completed_appointments, key=lambda item: item.appointment_date, default=None)
        next_appointment = min(upcoming_appointments, key=lambda item: item.appointment_date, default=None)
        rows.append(
            {
                "profile": patient,
                "age": calculate_age(patient.date_of_birth),
                "last_visit": last_visit,
                "next_appointment": next_appointment,
                "status": "Active" if patient.user.is_active else "Inactive",
            }
        )
    return rows


def patient_kpis(patients):
    today = timezone.localdate()
    month_start = today.replace(day=1)
    all_patients = PatientProfile.objects.select_related("user")
    return [
        {"label": "Total Patients", "value": all_patients.count(), "change": "+10%", "icon": "groups"},
        {"label": "New Patients This Month", "value": all_patients.filter(user__date_joined__date__gte=month_start).count(), "change": "+6%", "icon": "person_add"},
        {"label": "Active Patients", "value": all_patients.filter(user__is_active=True).count(), "change": "+12%", "icon": "verified_user"},
        {"label": "Patients With Appointments", "value": all_patients.annotate(total=Count("appointments")).filter(total__gt=0).count(), "change": "+8%", "icon": "event_available"},
        ]


def _dashboard_base_ctx(user):
    from notifications.models import Notification
    if user.role == User.Role.RECEPTIONIST:
        return {
            "base_template": "dashboard/receptionist_base.html",
            "unread_count": Notification.objects.filter(user=user, is_read=False).count(),
        }
    return {"base_template": "dashboard/base.html"}


@role_required(User.Role.ADMIN, User.Role.DENTIST, User.Role.RECEPTIONIST)
def patient_list(request):
    patient_form = PatientForm()
    modal_open = False

    if request.method == "POST":
        if request.user.role not in {User.Role.ADMIN, User.Role.RECEPTIONIST} and not request.user.is_superuser:
            messages.error(request, "You do not have permission to add patients.")
            return redirect("patients:list")
        patient_form = PatientForm(request.POST)
        modal_open = True
        if patient_form.is_valid():
            patient_form.save()
            messages.success(request, "Patient created successfully.")
            return redirect("patients:list")
        messages.error(request, "Please correct the patient form errors.")

    patients = base_patient_queryset(request)
    ctx = _dashboard_base_ctx(request.user)
    ctx.update({
        "kpis": patient_kpis(patients),
        "patient_rows": build_patient_rows(patients),
        "patient_form": patient_form,
        "modal_open": modal_open,
        "gender_choices": PatientProfile.Gender.choices,
        "appointment_status_choices": Appointment.Status.choices,
        "filters": {
            "q": request.GET.get("q", ""),
            "gender": request.GET.get("gender", ""),
            "status": request.GET.get("status", ""),
            "appointment_status": request.GET.get("appointment_status", ""),
        },
    })
    return render(request, "patients/patient_list.html", ctx)


@role_required(User.Role.ADMIN, User.Role.RECEPTIONIST)
def patient_create(request):
    form = PatientForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Patient created successfully.")
        return redirect("patients:list")
    return render(request, "patients/patient_form.html", {"form": form, "title": "Add New Patient"})


@role_required(User.Role.ADMIN, User.Role.DENTIST, User.Role.RECEPTIONIST)
def patient_detail(request, pk):
    patient = get_object_or_404(PatientProfile.objects.select_related("user"), pk=pk)
    treatments = Treatment.objects.filter(patient=patient).select_related("dentist__user")[:8]
    appointments = Appointment.objects.filter(patient=patient).select_related("dentist__user")[:8]
    return render(
        request,
        "patients/patient_detail.html",
        {"patient": patient, "age": calculate_age(patient.date_of_birth), "treatments": treatments, "appointments": appointments},
    )


@role_required(User.Role.ADMIN, User.Role.RECEPTIONIST)
def patient_edit(request, pk):
    patient = get_object_or_404(PatientProfile.objects.select_related("user"), pk=pk)
    form = PatientForm(request.POST or None, instance=patient)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Patient updated successfully.")
        return redirect("patients:list")
    return render(request, "patients/patient_form.html", {"form": form, "title": "Edit Patient"})


def patient_payload(patient):
    return {
        "id": patient.pk,
        "full_name": patient.user.full_name,
        "email": patient.user.email,
        "phone": patient.user.phone,
        "gender": patient.gender,
        "date_of_birth": patient.date_of_birth.isoformat() if patient.date_of_birth else "",
        "address": patient.address,
        "emergency_contact": patient.emergency_contact,
        "medical_history": patient.medical_history,
        "allergies": patient.allergies,
        "is_active": patient.user.is_active,
        "display": {
            "name": str(patient),
            "email": patient.user.email,
            "phone": patient.user.phone or "-",
            "gender": patient.get_gender_display() or "-",
            "status": "Active" if patient.user.is_active else "Inactive",
        },
    }


@role_required(User.Role.ADMIN, User.Role.DENTIST, User.Role.RECEPTIONIST)
def patient_json(request, pk):
    patient = get_object_or_404(PatientProfile.objects.select_related("user"), pk=pk)
    return JsonResponse({"ok": True, "record": patient_payload(patient)})


@require_POST
@role_required(User.Role.ADMIN, User.Role.RECEPTIONIST)
def patient_update(request, pk):
    patient = get_object_or_404(PatientProfile.objects.select_related("user"), pk=pk)
    form = PatientForm(request.POST, instance=patient)
    if form.is_valid():
        patient = form.save()
        return JsonResponse({"ok": True, "message": "Patient updated successfully.", "record": patient_payload(patient)})
    return JsonResponse({"ok": False, "message": "Please correct the patient form errors.", "errors": form.errors}, status=400)


@require_POST
@role_required(User.Role.ADMIN, User.Role.RECEPTIONIST)
def patient_delete(request, pk):
    patient = get_object_or_404(PatientProfile.objects.select_related("user"), pk=pk)
    patient.user.delete()
    messages.success(request, "Patient deleted successfully.")
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": True, "message": "Patient deleted successfully."})
    return redirect("patients:list")


@role_required(User.Role.ADMIN, User.Role.RECEPTIONIST)
def export_patients_pdf(request):
    rows = build_patient_rows(base_patient_queryset(request))
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="patients.pdf"'

    buffer = BytesIO()
    document = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    data = [["Name", "Phone", "Email", "Gender", "Age", "Status"]]
    for row in rows:
        patient = row["profile"]
        data.append(
            [
                str(patient),
                patient.user.phone or "-",
                patient.user.email,
                patient.get_gender_display() or "-",
                row["age"],
                row["status"],
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16a34a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )
    document.build([Paragraph("DentalCare Patient List", styles["Title"]), Spacer(1, 12), table])
    response.write(buffer.getvalue())
    buffer.close()
    return response


@role_required(User.Role.ADMIN, User.Role.RECEPTIONIST)
def export_patients_excel(request):
    rows = build_patient_rows(base_patient_queryset(request))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Patients"
    headers = ["Name", "Phone", "Email", "Gender", "Age", "Status"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="16A34A")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for row in rows:
        patient = row["profile"]
        sheet.append(
            [
                str(patient),
                patient.user.phone or "-",
                patient.user.email,
                patient.get_gender_display() or "-",
                row["age"],
                row["status"],
            ]
        )

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 3

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="patients.xlsx"'
    workbook.save(response)
    return response
