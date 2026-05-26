import json
from datetime import date
from io import BytesIO

from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Spacer, Table

from accounts.models import User
from accounts.permissions import role_required
from appointments.models import Appointment
from billing.models import Invoice
from dentists.models import DentistProfile
from patients.models import PatientProfile
from treatments.models import Treatment


def _apply_filters(request):
    appointments = Appointment.objects.select_related("patient__user", "dentist__user")
    invoices = Invoice.objects.select_related("patient__user")
    treatments = Treatment.objects.select_related("patient__user", "dentist__user")

    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    dentist_id = request.GET.get("dentist", "").strip()
    status = request.GET.get("status", "").strip()

    if date_from:
        appointments = appointments.filter(appointment_date__gte=date_from)
        invoices = invoices.filter(created_at__date__gte=date_from)
        treatments = treatments.filter(treatment_date__gte=date_from)
    if date_to:
        appointments = appointments.filter(appointment_date__lte=date_to)
        invoices = invoices.filter(created_at__date__lte=date_to)
        treatments = treatments.filter(treatment_date__lte=date_to)
    if dentist_id:
        appointments = appointments.filter(dentist_id=dentist_id)
        treatments = treatments.filter(dentist_id=dentist_id)
    if status:
        appointments = appointments.filter(status=status)

    return appointments, invoices, treatments


@role_required(User.Role.ADMIN)
def report_list(request):
    appointments, invoices, treatments = _apply_filters(request)

    paid_revenue = invoices.filter(status=Invoice.Status.PAID).aggregate(t=Sum("amount"))["t"] or 0
    kpis = [
        {"label": "Total Patients", "value": PatientProfile.objects.count(), "icon": "groups"},
        {"label": "Total Appointments", "value": appointments.count(), "icon": "calendar_month"},
        {"label": "Total Revenue", "value": f"${paid_revenue:,.2f}", "icon": "paid"},
        {"label": "Completed Treatments", "value": treatments.count(), "icon": "healing"},
    ]

    # Monthly chart data – all-time last 12 months (not filtered)
    today = timezone.localdate()
    month_labels, month_appts, month_revenue = [], [], []
    for i in range(11, -1, -1):
        month = today.month - i
        year = today.year
        while month <= 0:
            month += 12
            year -= 1
        month_labels.append(date(year, month, 1).strftime("%b %Y"))
        month_appts.append(
            Appointment.objects.filter(appointment_date__year=year, appointment_date__month=month).count()
        )
        month_revenue.append(
            float(
                Invoice.objects.filter(
                    status=Invoice.Status.PAID,
                    created_at__year=year,
                    created_at__month=month,
                ).aggregate(t=Sum("amount"))["t"] or 0
            )
        )

    status_data = {
        "Pending": Appointment.objects.filter(status=Appointment.Status.PENDING).count(),
        "Approved": Appointment.objects.filter(status=Appointment.Status.APPROVED).count(),
        "Completed": Appointment.objects.filter(status=Appointment.Status.COMPLETED).count(),
        "Cancelled": Appointment.objects.filter(status=Appointment.Status.CANCELLED).count(),
    }

    treatment_dentists = (
        Treatment.objects.values("dentist__user__first_name", "dentist__user__last_name")
        .annotate(count=Count("id"))
        .order_by("-count")[:6]
    )
    treatment_labels = [
        f"{d['dentist__user__first_name']} {d['dentist__user__last_name']}" for d in treatment_dentists
    ]
    treatment_counts = [d["count"] for d in treatment_dentists]

    context = {
        "kpis": kpis,
        "appointments": appointments[:50],
        "invoices": invoices[:50],
        "treatments": treatments[:50],
        "dentists": DentistProfile.objects.select_related("user").all(),
        "status_choices": Appointment.Status.choices,
        "month_labels": json.dumps(month_labels),
        "month_appts": json.dumps(month_appts),
        "month_revenue": json.dumps(month_revenue),
        "status_labels": json.dumps(list(status_data.keys())),
        "status_values": json.dumps(list(status_data.values())),
        "treatment_labels": json.dumps(treatment_labels),
        "treatment_counts": json.dumps(treatment_counts),
    }
    return render(request, "reports/report_list.html", context)


def _appointment_rows(qs):
    rows = [["Patient", "Dentist", "Date", "Time", "Status"]]
    for a in qs:
        rows.append([str(a.patient), str(a.dentist), str(a.appointment_date), str(a.appointment_time), a.get_status_display()])
    return rows


def _invoice_rows(qs):
    rows = [["Invoice No", "Patient", "Amount", "Status", "Date"]]
    for i in qs:
        rows.append([f"INV-{i.pk:05d}", str(i.patient), str(i.amount), i.get_status_display(), str(i.created_at.date())])
    return rows


def _treatment_rows(qs):
    rows = [["Patient", "Dentist", "Diagnosis", "Treatment Date"]]
    for t in qs:
        rows.append([str(t.patient), str(t.dentist), t.diagnosis[:100], str(t.treatment_date)])
    return rows


@role_required(User.Role.ADMIN)
def export_pdf(request):
    appointments, invoices, treatments = _apply_filters(request)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    doc.build([
        Table(_appointment_rows(appointments)),
        Spacer(1, 0.3 * inch),
        Table(_invoice_rows(invoices)),
        Spacer(1, 0.3 * inch),
        Table(_treatment_rows(treatments)),
    ])
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dental_report.pdf"'
    response.write(buffer.getvalue())
    return response


@role_required(User.Role.ADMIN)
def export_excel(request):
    appointments, invoices, treatments = _apply_filters(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Appointments"
    for row in _appointment_rows(appointments):
        ws.append(row)
    ws2 = wb.create_sheet("Revenue")
    for row in _invoice_rows(invoices):
        ws2.append(row)
    ws3 = wb.create_sheet("Treatments")
    for row in _treatment_rows(treatments):
        ws3.append(row)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="dental_report.xlsx"'
    wb.save(response)
    return response
